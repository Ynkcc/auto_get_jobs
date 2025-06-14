# 标准库导入
import hashlib
import json
import logging
import os
import re
import threading
import time
from typing import List, Dict
import itertools
from urllib.parse import urlencode

# 第三方库导入
import asyncio
import aiohttp
import pandas as pd
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
import mimetypes
import yaml
from playwright.async_api import async_playwright, Page
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import sqlite3

# 本地模块导入
from utils.session_manager import SessionManager
logger = logging.getLogger(__name__)

# 定义基础URL
BASE_URL = 'https://www.zhipin.com'


def save_jobs_to_csv(jobs: List[Dict], filename: str = 'jobs.csv') -> None:
    """
    将 jobs 数据导出为 CSV 文件
    :param jobs: 岗位数据列表，每个岗位是一个字典
    :param filename: 导出的文件名，默认为 jobs.csv
    """
    try:
        df = pd.DataFrame(jobs)
        df.to_csv(filename, index=False, encoding='utf-8')
        logger.info(f"数据已成功导出到 {filename}")
    except Exception as e:
        logger.error(f"导出 CSV 文件时发生错误：{e}")

def export_to_xlsx(db_path, export_dir):
    """
    将 SQLite 数据库中的 job_details 表导出为 XLSX 文件，并应用筛选和格式化。

    Args:
        db_path (str): SQLite 数据库文件路径。
        export_dir (str): 要保存 XLSX 文件的目录路径。
    """
    timestamp = datetime.datetime.now().strftime("%H%M")
    filename = f"jobs_{timestamp}.xlsx"
    xlsx_path = os.path.join(export_dir, filename)

    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM job_details ORDER BY first_added_time DESC", conn)
    df.to_excel(xlsx_path, index=False, engine='openpyxl')
    conn.close()

    workbook = load_workbook(xlsx_path)
    worksheet = workbook.active
    cols_to_format = ['postDescription', 'analysis_think']
    target_width = 40
    for col_idx, column_cell in enumerate(worksheet[1], 1):
        column_letter = get_column_letter(col_idx)
        if column_cell.value in cols_to_format:
            worksheet.column_dimensions[column_letter].width = target_width
            logger.info(f"设置列 '{column_cell.value}' ({column_letter}) 宽度为 {target_width}")
    # 创建居中对齐样式,以及自动化换行
    center_alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    # 遍历工作表中所有行和单元格
    for row in worksheet.iter_rows():  # 遍历每一行
        for cell in row:        # 遍历行中的每个单元格
            cell.alignment = center_alignment
    workbook.save(xlsx_path)
    logger.info(f"数据已成功导出并格式化到 {xlsx_path}")



class BrowserSessionHandler:
    def __init__(self, page: Page, login_data_file: str, loop: asyncio.AbstractEventLoop):
        self.page = page
        self.login_data_file = login_data_file
        self.save_task = None
        self.loop = loop

    async def load(self) -> bool:
        """加载登录数据"""
        try:
            with open(self.login_data_file, "r", encoding="utf-8") as f:
                login_data = json.load(f)

            await self.page.context.clear_cookies()
            if login_data.get("cookies"):
                await self.page.context.add_cookies(login_data["cookies"])

            await self.page.evaluate("() => localStorage.clear()")
            if login_data.get("localStorage"):
                for k, v in login_data["localStorage"].items():
                    await self.page.evaluate(f"() => localStorage.setItem('{k}', '{v}')")

            await self.page.evaluate("() => sessionStorage.clear()")
            if login_data.get("sessionStorage"):
                for k, v in login_data["sessionStorage"].items():
                    await self.page.evaluate(f"() => sessionStorage.setItem('{k}', '{v}')")

            logger.info("登录数据已加载")
            return True
        except FileNotFoundError:
            logger.warning("登录文件不存在")
            return False
        except Exception as e:
            logger.error(f"加载失败: {str(e)}")
            return False

    async def save(self) -> bool:
        """保存登录数据"""
        try:
            directory = os.path.dirname(self.login_data_file)
            os.makedirs(directory, exist_ok=True)

            cookies = await self.page.context.cookies()
            local_storage = await self.page.evaluate("() => Object.assign({}, localStorage)")
            session_storage = await self.page.evaluate("() => Object.assign({}, sessionStorage)")

            login_data = {
                "cookies": cookies,
                "localStorage": local_storage,
                "sessionStorage": session_storage
            }

            with open(self.login_data_file, "w", encoding="utf-8") as f:
                json.dump(login_data, f, indent=2, ensure_ascii=False)

            logger.info("登录数据已保存")
            return True
        except Exception as e:
            logger.error(f"保存失败: {str(e)}")
            return False

    async def start_autosave(self, interval=60):
        """启动定时保存任务"""
        async def saver():
            while True:
                await self.save()
                await asyncio.sleep(interval)

        self.save_task = asyncio.create_task(saver())
        logger.info("自动保存已启动")

    async def stop_autosave(self):
        """停止定时保存任务"""
        if self.save_task:
            self.save_task.cancel()
            try:
                await self.save_task
            except asyncio.CancelledError:
                pass
            logger.info("自动保存已停止")

    async def clear_data(self) -> bool:
        """清除浏览器数据"""
        try:
            await self.page.context.clear_cookies()
            await self.page.evaluate("() => localStorage.clear()")
            await self.page.evaluate("() => sessionStorage.clear()")
            logger.info("浏览器数据已清空，可重新登录")
            return True
        except Exception as e:
            logger.error(f"清除数据失败: {str(e)}")
            return False


async def init_driver(playwright_config):
    """
    初始化 Playwright 浏览器
    :param playwright_config: playwright配置对象
    :return: Playwright 浏览器实例
    """
    browser_type = playwright_config.browser_type.lower()

    try:
        playwright = await async_playwright().start()

        if browser_type == "chromium":
            browser = await playwright.chromium.launch(
                headless=playwright_config.headless,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    '--log-level=3',  # 只记录严重错误
                ],
                ignore_default_args=["--enable-automation"],
            )
        elif browser_type == "firefox":
            browser = await playwright.firefox.launch(
                headless=playwright_config.headless,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    '--log-level=3',  # 只记录严重错误
                ],
                ignore_default_args=["--enable-automation"],
            )
        elif browser_type == "webkit":
            browser = await playwright.webkit.launch(
                headless=playwright_config.headless,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    '--log-level=3',  # 只记录严重错误
                ],
                ignore_default_args=["--enable-automation"],
            )
        elif browser_type == "edge":
            browser = await playwright.chromium.launch(
                channel="msedge",
                headless=playwright_config.headless,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    '--log-level=3',  # 只记录严重错误
                ],
                ignore_default_args=["--enable-automation"],
            )
        elif browser_type == "chrome":
            browser = await playwright.chromium.launch(
                channel="chrome",
                headless=playwright_config.headless,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    '--log-level=3',  # 只记录严重错误
                ],
                ignore_default_args=["--enable-automation"],
            )
        else:
            raise ValueError(f"不支持的浏览器类型: {browser_type}")

        context = await browser.new_context()
        page = await context.new_page()
        return page

    except Exception as e:
        logger.error(f"浏览器初始化失败: {str(e)}")
        raise

def build_search_url(job_search):
    """
    构造搜索URL（支持参数合并配置）
    """
    # 加载区域代码数据
    with open('config/search_params_config.json', 'r', encoding='utf-8') as f:
        params_data = json.load(f)

    location_dicts = {}
    city_code_map = params_data["cityCode"]

    # 处理 areas 配置（优先级高）
    if job_search.areas:
        for city_name, districts in job_search.areas.items():
            if city_name not in city_code_map:
                logger.warning(f"未找到城市 [{city_name}] 的编码，已跳过")
                continue

            city_entry = city_code_map[city_name]
            city_code = list(city_entry.keys())[0]

            valid_districts = [
                city_entry[city_code][district]
                for district in districts
                if district in city_entry[city_code]
            ]
            
            invalid = set(districts) - set(city_entry[city_code].keys())
            if invalid:
                logger.warning(f"城市 [{city_name}] 下未找到区域: {', '.join(invalid)}")

            if valid_districts:
                location_dicts.setdefault(city_code, []).extend(valid_districts)
    else:
        # 处理 city 配置
        for city_name in job_search.city.values:
            if city_name not in city_code_map:
                logger.warning(f"未找到城市 [{city_name}] 的编码，已跳过")
                continue

            city_entry = city_code_map[city_name]
            city_code = list(city_entry.keys())[0]

            # 根据 expand_to_district 决定是否添加区域
            if job_search.city.expand_to_district:
                location_dicts[city_code] = list(city_entry[city_code].values())
            else:
                location_dicts[city_code] = [] # 不展开到区域，则区域列表为空

    if not location_dicts:
        raise ValueError("未找到有效的城市/区域配置")

    base_url = "https://www.zhipin.com/web/geek/job"

    def process_filter(filter_config, param_map: dict) -> List[str]:
        """处理过滤参数合并逻辑"""
        codes = [str(param_map[v]) for v in filter_config.values]
        return [','.join(codes)] if filter_config.combine and codes else codes

    # 构建参数配置
    params_config = {
        'degree': process_filter(job_search.degree, params_data["degree"]),
        'position': process_filter(job_search.position, params_data["position"]),
        'industry': process_filter(job_search.industry, params_data["industry"]),
        'experience': process_filter(job_search.experience, params_data["experience"]),
        'scale': process_filter(job_search.scale, params_data["scale"]),
        'stage': process_filter(job_search.stage, params_data["stage"]),
        'salary': [str(params_data["salary"][v]) for v in job_search.salary],
        'jobType': [str(params_data["jobType"][v]) for v in job_search.jobType],
        'query': job_search.query
    }
    params_config = {k: v for k, v in params_config.items() if v}

    # -- 修改开始 --
    base_params_list = []
    for city_code, districts in location_dicts.items():
        # 如果区域列表为空（包括 areas 为空或 city.expand_to_district 为 false 的情况）
        # 则只搜索城市
        if not districts:
            base_params_list.append({'city': city_code})
            continue

        # 如果区域列表不为空，则遍历所有区域
        for district_code in districts:
            base_params_list.append({'city': city_code, 'multiBusinessDistrict': district_code})
    # -- 修改结束 --
    
    url_list = []
    if not params_config:
        for params in base_params_list:
            param_str = '&'.join(f"{k}={v}" for k, v in params.items())
            url_list.append(f"{base_url}?{param_str}")
        return url_list
    
    param_keys = list(params_config.keys())
    param_combinations = list(itertools.product(*params_config.values()))

    for base_param in base_params_list:
        for combination in param_combinations:
            merged_params = base_param.copy()
            merged_params.update(zip(param_keys, combination))
            param_str = urlencode(merged_params)
            url_list.append(f"{base_url}?{param_str}")

    return url_list


async def get_page_jobs_info(page: Page):
    jobs = []
    job_cards = page.locator('li.job-card-wrapper')
    count = await job_cards.count()

    for i in range(count):
        job_card = job_cards.nth(i)
        job_name = await job_card.locator('span.job-name').inner_text()
        job_salary = await job_card.locator('span.salary').inner_text()
        job_link = await job_card.locator('a.job-card-left').get_attribute('href')
        company_name = await job_card.locator('h3.company-name').inner_text()
        company_tags = []
        company_tag_list = job_card.locator('ul.company-tag-list')
        if await company_tag_list.count() > 0:
            tag_elements = await company_tag_list.locator('li').all_inner_texts()
            company_tags = tag_elements
        jobs.append({
            'job_name': job_name,
            'job_salary': job_salary,
            'job_link': job_link,
            'company_name': company_name,
            'company_tags': company_tags
        })
    return jobs


async def next_page(page: Page):
    try:
        # 使用 Playwright 的选择器
        next_page_button = page.locator(".ui-icon-arrow-right")
        print("开始采集下一页")
        # 判断按钮是否禁用
        if "disabled" in (await next_page_button.locator("xpath=..").get_attribute("class")):
            print("没有下一页了，退出循环。")
            return False
        # 点击下一页按钮
        await next_page_button.click()
        return True
    except Exception:
        print("可能找不到下一页，退出循环。")
        return False


def filter_jobs_by_salary(jobs: List[Dict], min_expected_salary: float, max_expected_salary: float) -> List[Dict]:
    """
    根据期望薪资范围过滤岗位
    :param jobs: 岗位列表
    :param min_expected_salary: 期望最低薪资（单位：K）
    :param max_expected_salary: 期望最高薪资（单位：K）
    :return: 符合条件的岗位列表
    """
    jobs_matching_salary = []

    for job in jobs:
        job_name = job['job_name']
        job_salary = job['job_salary']
        original_salary = job_salary  # 保留原始值用于日志
        try:
            # 预处理：去除类似"·13薪"的后缀
            job_salary = re.sub(r'·\d+薪', '', job_salary)

            # 统一转换为小写方便处理
            salary_str = job_salary.lower()

            if '元/天' in salary_str:
                # 日薪处理（按22工作日/月）
                daily = salary_str.replace('元/天', '')
                daily_range = [float(x) for x in daily.split('-')]
                min_d = daily_range[0]
                max_d = daily_range[-1]  # 处理单值和范围
                min_monthly = min_d * 22 / 1000
                max_monthly = max_d * 22 / 1000

            elif 'k' in salary_str:
                # K表示的月薪
                k_str = salary_str.replace('k', '')
                k_range = [float(x) for x in k_str.split('-')]
                min_monthly = k_range[0]
                max_monthly = k_range[-1]

            elif '元/月' in salary_str:
                # 直接月薪
                monthly = salary_str.replace('元/月', '')
                monthly_range = [float(x) for x in monthly.split('-')]
                min_monthly = monthly_range[0] / 1000
                max_monthly = monthly_range[-1] / 1000

            elif '元/周' in salary_str:
                # 周薪处理（按4周/月）
                weekly = salary_str.replace('元/周', '')
                weekly_range = [float(x) for x in weekly.split('-')]
                min_monthly = weekly_range[0] * 4 / 1000
                max_monthly = weekly_range[-1] * 4 / 1000

            elif '元/时' in salary_str:
                # 时薪处理（按8小时/天，22天/月）
                hourly = salary_str.replace('元/时', '')
                hourly_range = [float(x) for x in hourly.split('-')]
                min_monthly = hourly_range[0] * 8 * 22 / 1000
                max_monthly = hourly_range[-1] * 8 * 22 / 1000

            else:
                logger.warning(f"未知薪资格式 | 岗位: {job_name} | 薪资: {original_salary}")
                continue
        except (ValueError, IndexError) as e:
            logger.warning(f"薪资解析失败 | 岗位: {job_name} | 原始值: {original_salary} | 错误: {str(e)}")
            continue
        except Exception as e:
            logger.error(f"意外错误处理薪资 | 岗位: {job_name} | 错误: {str(e)}")
            continue

        # 薪资判断逻辑
        if min_expected_salary <= min_monthly and max_monthly <= max_expected_salary:
            jobs_matching_salary.append(job)
            logger.debug(f"符合条件 | 岗位: {job_name} | 范围: {min_monthly:.1f}-{max_monthly:.1f}K")
        else:
            logger.info(f"薪资不符合 | 岗位: {job_name} | 当前范围: {min_monthly:.1f}-{max_monthly:.1f}K | 期望范围: {min_expected_salary:.1f}-{max_expected_salary:.1f}K")

    return jobs_matching_salary


def parse_params(link):
    """
    从招聘链接中提取关键参数
    """
    pattern = r'/job_detail/([^.]+)\.html\?lid=([^&]+)&securityId=([^&]+)'
    match = re.search(pattern, link)
    return match.groups() if match else None


def get_user_info():
    url = "https://www.zhipin.com/wapi/zpuser/wap/getUserInfo.json"
    session = SessionManager.get_sync_session()
    try:
        user_info = session.get(url).json()
        user_id = user_info['zpData']['userId']
        user_name = user_info['zpData']['name']
        true_man = user_info['zpData']['trueMan']
        print(f"成功获取到用户信息: 用户名是：{user_name},账号id是：{user_id}")
        if not true_man:
            print("警告：本账号已被BOSS直聘标记")
    except Exception as e:
        print("获取用户信息失败")
        print(e)
        return "UNKNOWN", None
    return user_id, user_info

def get_wt2():
    """获取wt2验证参数"""
    try:
        url = "https://www.zhipin.com/wapi/zppassport/get/wt"
        session = SessionManager.get_sync_session()
        response = session.get(url, timeout=10)
        response.raise_for_status()
        wt2_data = response.json()

        if wt2_data['code'] == 0:
            return wt2_data['zpData'].get('wt2')
        raise Exception(f"获取wt2失败: {wt2_data.get('message')}")
    except Exception as e:
        logger.error(f"获取wt2异常: {str(e)}")
        return None


def calculate_md5(file_path):
    """计算文件的 MD5 哈希值"""
    hash_md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def full_upload_image(file_path,securityId):
    """
    上传图片
    """
    try:
        url = "https://www.zhipin.com/wapi/zpupload/image/uploadSingle"
        session = SessionManager.get_sync_session()
        mime_type, _ = mimetypes.guess_type(file_path)
        with open(file_path, "rb") as f:
            mp_encoder = MultipartEncoder(
                fields=[
                    ('securityId', securityId),
                    ('source', 'chat_file'),
                    ('file', (os.path.basename(file_path), f, mime_type ))
                ]
            )
            headers = {'Content-Type': mp_encoder.content_type,
            'user-agent':"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
            }
            response = session.post(url, data=mp_encoder,headers=headers)
        zp_data = response.json()["zpData"]
        origin_width = zp_data["metadata"]["width"]
        origin_height = zp_data["metadata"]['height']
        tiny_width = 200
        tiny_height = int(tiny_width * origin_height / origin_width)
        result = {
            "tinyImage": {
                "url": zp_data['tinyUrl'],
                "width": tiny_width,
                "height": tiny_height
            },
            "originImage": {
                "url": zp_data['url'],
                "width": origin_width,
                "height": origin_height
            }
        }
        return result
    except Exception:
        logger.exception(f"上传简历图片出现错误")
        return None

def quickly_upload_image(file_md5, securityId):
    # 快速上传接口（若服务端已有相同文件则直接返回结果）
    url = "https://www.zhipin.com/wapi/zpupload/quicklyUpload"
    data = {
        "fileMd5": file_md5,  # use file_md5 instead of calculate_md5(file_path)
        "fileSize": 0,  # remove os.path.getsize(file_path),
        "source": "chat_file",
        "securityId": securityId
    }
    session = SessionManager.get_sync_session()
    response = session.post(url, data=data)
    zp_data = response.json()["zpData"]
    if zp_data.get("url"):
        origin_width = zp_data["metadata"]["width"]
        origin_height = zp_data["metadata"]['height']
        tiny_width = 200
        tiny_height = int(tiny_width * origin_height / origin_width)
        result = {
            "tinyImage": {
                "url": zp_data['tinyUrl'],
                "width": tiny_width,
                "height": tiny_height
            },
            "originImage": {
                "url": zp_data['url'],
                "width": origin_width,
                "height": origin_height
            }
        }
    else:
        result = False
    return result


def upload_image(file_path, securityId, resume_image_md5=None): 
    quickly_upload_result = quickly_upload_image(resume_image_md5, securityId)
    if quickly_upload_result:
        return quickly_upload_result
    else:
        full_upload_result =  full_upload_image(file_path, securityId) 
        return full_upload_result
    



class TokenBucket:
    def __init__(self, rate: int, capacity: int):
        self.rate = rate  # 每秒生成的令牌数
        self.capacity = capacity  # 桶的容量
        self.tokens = capacity  # 初始化令牌数为桶容量
        self.last_check = time.time()

    async def get_token(self):
        while True:
            now = time.time()
            elapsed = now - self.last_check
            if elapsed > 1:
                # 每秒重新生成令牌
                self.tokens = min(self.capacity, self.tokens + int(elapsed * self.rate))
                self.last_check = now
            if self.tokens > 0:
                self.tokens -= 1
                break
            else:
                # 等待直到有令牌
                await asyncio.sleep(1 / self.rate)

async def get_job_info(securityId, lid, max_retries=3):
    path = "/wapi/zpgeek/job/card.json"
    url = f"{BASE_URL}{path}"
    params = {
        "securityId": securityId,
        "lid": lid
    }

    async def _request(attempt):
        try:
            session = await SessionManager.get_async_session()
            async with session.get(
                url,
                params=params,
                timeout=aiohttp.ClientTimeout(total=30 + attempt*5)  # 动态超时
            ) as response:
                if response.status == 200:
                    return await response.json()
                elif response.status == 403:
                    logger.warning(f"[403 Forbidden] 可能触发反爬虫 securityId={securityId}")
                    return None
                else:
                    logger.error(f"请求失败 HTTP {response.status} - {await response.text()}")
                    return None
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logger.warning(f"请求异常 [{type(e).__name__}] 第{attempt}次重试: {str(e)}")
            return None

    # 重试逻辑
    for attempt in range(1, max_retries + 1):
        result = await _request(attempt)
        if result is not None:
            return result
        await asyncio.sleep(min(2 ** attempt, 10))  # 指数退避

    logger.error(f"请求失败（已达最大重试次数 {max_retries}）")
    return None

async def start_chat(securityId, jobId, lid):
    path = "/wapi/zpgeek/friend/add.json"
    url = f"{BASE_URL}{path}"
    params = {
        "securityId": securityId,
        "jobId": jobId,
        "lid": lid
    }

    session = await SessionManager.get_async_session()
    async with session.get(url, params=params, timeout=30) as response:
        if response.status == 200:
            return await response.json()
        else:
            response.raise_for_status()


# 以下接口待验证
async def get_boss_list(page: int, cookies=None, headers=None):
    path = "/wapi/zprelation/friend/getGeekFriendList.json"
    url = f"https://www.zhipin.com{path}"
    params = {"page": page}

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, cookies=cookies, headers=headers) as response:
            return await response.json()


async def get_boss_data(encryptBossId: str, securityId: str, source_type="0", cookies=None, headers=None):
    path = "/wapi/zpgeek/chat/bossdata.json"
    url = f"https://www.zhipin.com{path}"
    params = {
        "bossId": encryptBossId,
        "bossSource": source_type,
        "securityId": securityId
    }

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, cookies=cookies, headers=headers) as response:
            return await response.json()


async def get_history_msg(encryptBossId: str, securityId: str, cookies=None, headers=None):
    path = "/wapi/zpchat/geek/historyMsg"
    url = f"https://www.zhipin.com{path}"
    params = {
        "bossId": encryptBossId,
        "groupId": encryptBossId,
        "securityId": securityId,
        "maxMsgId": "0",
        "c": "20",
        "page": "1"
    }

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, cookies=cookies, headers=headers) as response:
            return await response.json()


async def request_send_resume(bossId: str, resumeId: str, cookies=None, headers=None):
    path = "/geek/new/requestSendResume.json"
    url = f"https://www.zhipin.com{path}"
    params = {
        "bossId": bossId,
        "resumeId": resumeId,
        "toSource": "0"
    }

    async with aiohttp.ClientSession() as session:
        async with session.get(url, params=params, cookies=cookies, headers=headers) as response:
            return await response.json()


async def get_resumes(cookies=None, headers=None):
    path = "/wapi/zpgeek/resume/attachment/checkbox.json"
    url = f"https://www.zhipin.com{path}"

    async with aiohttp.ClientSession() as session:
        async with session.get(url, cookies=cookies, headers=headers) as response:
            return await response.json()